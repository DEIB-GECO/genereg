# coding: utf-8

# Import libraries
import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
import pickle
import numpy as np


def summarize_reg(gene_set, n_data_matrix):

	"""
	The SUMMARIZE_REG operation summarizes all the data analysis results, by collecting them in convenient tables that exported locally in Excel files.

	:param gene_set: the set of genes of interest to summarize
	:param n_data_matrix: number identifying the data matrix to summarize (only 2,3 and 5 values are permitted)
	
	Example::
	
		import genereg as gr
		gr.SummaryResults.summarize_reg(gene_set='DNA_REPAIR', n_data_matrix=2)
		gr.SummaryResults.summarize_reg(gene_set='DNA_REPAIR', n_data_matrix=3)
		gr.SummaryResults.summarize_reg(gene_set='DNA_REPAIR', n_data_matrix=5)		
	"""

	# Check input parameters	
	if n_data_matrix not in [2, 3, 5]:
		raise ValueError('Data Matrix ERROR! Possible values: {2,3,5}')
	
	# Define the model to summarize
	model = str(n_data_matrix)
	
	# Define the previous model to check
	if model == '3':
		previous_model = str(int(model)-1)
	elif model == '5':
		previous_model = str(int(model)-2)

	# Import the dictionary of genes of interest with their candidate regulatory genes
	dict_RegulGenes = pickle.load(open('./2_Regulatory_Genes/dict_RegulGenes.p', 'rb'))

	# Import the list of genes of interest and extract in a list the Gene Symbols of all the genes belonging to the current gene set
	EntrezConversion_df = pd.read_excel('./Genes_of_Interest.xlsx',sheetname='Sheet1',header=0,converters={'GENE_SYMBOL':str,'ENTREZ_GENE_ID':str,'GENE_SET':str})

	SYMs_current_pathway = []
	for index, row in EntrezConversion_df.iterrows():
		sym = row['GENE_SYMBOL']
		path = row['GENE_SET']
		if path == gene_set:
			SYMs_current_pathway.append(sym)

	if (model == '3') or (model == '5'):
		# Create a list containing the Gene Symbols of the regulatory genes of the genes in the current gene set
		current_regulatory_genes = []
		for key, value in dict_RegulGenes.items():
			if key in SYMs_current_pathway:
				for gene in value:  
					if gene not in current_regulatory_genes:
						current_regulatory_genes.append(gene)
	
	if (model == '5'):
		# Create a list containing the Gene Symbols of genes in the other gene sets
		SYMs_other_pathways = []
		for index, row in EntrezConversion_df.iterrows():
			sym = row['GENE_SYMBOL']
			path = row['GENE_SET']
			if not (path == gene_set):
				SYMs_other_pathways.append(sym)
			
		# Create a list containing the Gene Symbols of the regulatory genes of the genes in the other gene sets
		regulatory_genes_other = []
		for key, value in dict_RegulGenes.items():
			if key not in SYMs_current_pathway:
				for gene in value:  
					if gene not in regulatory_genes_other:
						regulatory_genes_other.append(gene)   


	# Create a dataframe to store final summary results of feature selection and linear regression for each gene of interest
	if model == '2':
		lr_summary_df = pd.DataFrame(index=SYMs_current_pathway, columns=['Inital N° Features','Discarded Features','N° Features Selected','R2','Adj.R2']) 
	else:
		lr_summary_df = pd.DataFrame(index=SYMs_current_pathway, columns=['Inital N° Features','N° New Features w.r.t. Previous Model','Discarded Features','Features Available for Selection','N° Features Selected','R2','Adj.R2']) 

	for current_gene in SYMs_current_pathway:
	
		# Import the current and, if present, the previous model of the current gene
		gene_ID = EntrezConversion_df.loc[EntrezConversion_df['GENE_SYMBOL'] == current_gene, 'ENTREZ_GENE_ID'].iloc[0]
		model_gene_df = pd.read_excel('./4_Data_Matrix_Construction/Model'+model+'/Gene_'+gene_ID+'_['+current_gene+']'+'_('+gene_set+')-Model_v'+model+'.xlsx',sheetname='Sheet1',header=0)
		if not (model == '2'):
			previous_model_df = pd.read_excel('./4_Data_Matrix_Construction/Model'+previous_model+'/Gene_'+gene_ID+'_['+current_gene+']'+'_('+gene_set+')-Model_v'+previous_model+'.xlsx',sheetname='Sheet1',header=0)
		
		# Extract the list of new features, added to the current model, w.r.t. the previous one
		if not (model == '2'): 
			current_model_col_names = set(list(model_gene_df.columns.values))
			previous_model_col_names = set(list(previous_model_df.columns.values))
			new_features = list(current_model_col_names - previous_model_col_names)    
			lr_summary_df.set_value(current_gene,'N° New Features w.r.t. Previous Model',len(new_features))
	
		# Import the feature selection and linear regression summary tables
		feature_sel_df = pd.read_excel('./5_Data_Analysis/'+gene_set+'/FeatureSelection/M'+model+'/Feature_Selection_SUMMARY.xlsx',sheetname='Sheet1',header=0)
		lin_reg_df = pd.read_excel('./5_Data_Analysis/'+gene_set+'/LinearRegression/M'+model+'/Linear_Regression_R2_SCORES.xlsx',sheetname='Sheet1',header=0)
		
		# Extract and store the results in the summary dataframe
		n_features = feature_sel_df.get_value(current_gene,'TOT Inital N° Features')
		n_feat_discarded = feature_sel_df.get_value(current_gene,'Discarded Features')
		if not (model == '2'): 
			n_features_available = feature_sel_df.get_value(current_gene,'Features Available for Selection')
		n_feat_selected = feature_sel_df.get_value(current_gene,'N° Features Selected')
		lin_reg_r2_adj = lin_reg_df.get_value(current_gene,'Adj.R2')
		lin_reg_r2 = lin_reg_df.get_value(current_gene,'R2')
		lr_summary_df.set_value(current_gene,'Inital N° Features',n_features)
		lr_summary_df.set_value(current_gene,'Discarded Features',n_feat_discarded)
		if not (model == '2'): 
			lr_summary_df.set_value(current_gene,'Features Available for Selection',n_features_available)
		lr_summary_df.set_value(current_gene,'N° Features Selected',n_feat_selected)
		lr_summary_df.set_value(current_gene,'Adj.R2',lin_reg_r2_adj)
		lr_summary_df.set_value(current_gene,'R2',lin_reg_r2)

	# Export the summary dataframe in an Excel file
	lr_summary_df = lr_summary_df.sort_values(by=['Adj.R2'], ascending=[False])
	filename = './5_Data_Analysis/'+gene_set+'/Feature_Selection_and_Linear_Regression.xlsx'
	writer = ExcelWriter(filename,engine='openpyxl')
	try:
		writer.book = load_workbook(filename)
		writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
	except IOError:
		# if the file does not exist yet, I will create it
		pass
	lr_summary_df.to_excel(writer,'M'+model)
	writer.save()
		

	# Extract relevant features for each gene of the current gene set and store them in a summary table and define a dataframe to summarize the features selected for each model gene
	features_summary_df = pd.DataFrame(index=SYMs_current_pathway)

	for current_gene in SYMs_current_pathway:
		
		gene_ID = EntrezConversion_df.loc[EntrezConversion_df['GENE_SYMBOL'] == current_gene, 'ENTREZ_GENE_ID'].iloc[0]
		
		# Import the regression coefficients
		coeff_df = pd.read_excel('./5_Data_Analysis/'+gene_set+'/LinearRegression/M'+model+'/Coefficients/Coefficients_(M'+model+')-Gene_'+gene_ID+'_['+current_gene+'].xlsx',sheetname='Sheet1',header=0)

		# Import the confidence intervals
		ci_df = pd.read_excel('./5_Data_Analysis/'+gene_set+'/LinearRegression/M'+model+'/ConfidenceIntervals/Confidence_Intervals_(M'+model+')-Gene_'+gene_ID+'_['+current_gene+'].xlsx',sheetname='Sheet1',header=0)
		
		# Import the correlation matrix
		corr_df = pd.read_excel('./5_Data_Analysis/'+gene_set+'/LinearRegression/M'+model+'/CorrelationMatrix/Correlation_Matrix_(M'+model+')-Gene_'+gene_ID+'_['+current_gene+'].xlsx',sheetname='Sheet1',header=0)
		
		# Select the relevant features on the basis of the confidence intervals (i.e. if the confidence interval does not contain 0, then the feature is significant for the model)
		relevant_features = []
		for index, row in ci_df.iterrows():
			s = row['Significant Feature?']
			if s == 'YES':
				relevant_features.append(index)            
		
		# Create a dataframe to store the results and fill it with requested information
		relevant_features_df = pd.DataFrame(index=relevant_features, columns=['Regression Coefficient','Feature Description','Correlation with EXPRESSION ('+current_gene+')'])
		for index, row in coeff_df.iterrows():
			gene = row['feature']
			if gene in relevant_features:
				coeff = row['coefficient']
				relevant_features_df.set_value(gene,'Regression Coefficient',coeff)
		for index, row in corr_df.iterrows():
			if index in relevant_features:
				corr_with_target = row['EXPRESSION ('+current_gene+')']
				relevant_features_df.set_value(index,'Correlation with EXPRESSION ('+current_gene+')',corr_with_target)  
		
		# Add the features descriptions
		if model == '2':
			for f in relevant_features:
				if f in SYMs_current_pathway:
					descr = 'Gene of the '+gene_set+' set'
					relevant_features_df.set_value(f,'Feature Description',descr)
				elif 'METHYLATION' in f:
					descr = 'Methylation of the model gene ['+current_gene+'] in the '+gene_set+' set'
					relevant_features_df.set_value(f,'Feature Description',descr)
				elif f in dict_RegulGenes[current_gene]:
					descr = 'Candidate regulatory gene of the model gene ['+current_gene+'] of the '+gene_set+' set'
					relevant_features_df.set_value(f,'Feature Description',descr)
		
		elif model == '3':
			for f in relevant_features:
				if f in SYMs_current_pathway:
					descr = 'Gene of the '+gene_set+' set'
					relevant_features_df.set_value(f,'Feature Description',descr)
				elif 'METHYLATION' in f:
					descr = 'Methylation of the model gene ['+current_gene+'] in the '+gene_set+' set'
					relevant_features_df.set_value(f,'Feature Description',descr)
				elif f in dict_RegulGenes[current_gene]:
					descr = 'Candidate regulatory gene of the model gene ['+current_gene+'] of the '+gene_set+' set'
					relevant_features_df.set_value(f,'Feature Description',descr)
				elif not(f in dict_RegulGenes[current_gene]) and (f in current_regulatory_genes):
					descr = 'Candidate regulatory gene of the genes in the '+gene_set+' set'
					relevant_features_df.set_value(f,'Feature Description',descr)
		
		elif model == '5':
			for f in relevant_features:
				if f in SYMs_current_pathway:
					descr = 'Gene of the '+gene_set+' set'
					relevant_features_df.set_value(f,'Feature Description',descr)
				elif 'METHYLATION' in f:
					descr = 'Methylation of the model gene ['+current_gene+'] in the '+gene_set+' set'
					relevant_features_df.set_value(f,'Feature Description',descr)
				elif f in dict_RegulGenes[current_gene]:
					descr = 'Candidate regulatory gene of the model gene ['+current_gene+'] of the '+gene_set+' set'
					relevant_features_df.set_value(f,'Feature Description',descr)
				elif not(f in dict_RegulGenes[current_gene]) and (f in current_regulatory_genes):
					descr = 'Candidate regulatory gene of the genes in the '+gene_set+' set'
					relevant_features_df.set_value(f,'Feature Description',descr)
				elif f in SYMs_other_pathways:
					df_temp = EntrezConversion_df.loc[EntrezConversion_df['GENE_SYMBOL'] == f].copy()
					f_pathways = (df_temp.GENE_SET.unique()).tolist()
					descr = 'Gene of the gene sets: '+(', '.join(f_pathways))
					relevant_features_df.set_value(f,'Feature Description',descr)
				elif f in regulatory_genes_other:
					regulated_genes_other = []
					for key, value in dict_RegulGenes.items():
						if key in SYMs_other_pathways:
							if f in value:  
								regulated_genes_other.append(key)
					df_temp = EntrezConversion_df.loc[EntrezConversion_df['GENE_SYMBOL'].isin(regulated_genes_other)].copy()
					f_pathways = (df_temp.GENE_SET.unique()).tolist()
					descr = 'Candidate regulatory gene of the gene sets: '+(', '.join(f_pathways))
					relevant_features_df.set_value(f,'Feature Description',descr)           

				
		# Export the dataframe in an Excel file
		relevant_features_df = relevant_features_df.sort_values(by=['Regression Coefficient'], ascending=[False])
		filename = './5_Data_Analysis/'+gene_set+'/Relevant_Features-Gene_'+gene_ID+'_['+current_gene+'].xlsx'
		writer = ExcelWriter(filename,engine='openpyxl')
		try:
			writer.book = load_workbook(filename)
			writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
		except IOError:
			# if the file does not exist yet, I will create it
			pass
		relevant_features_df.to_excel(writer,'M'+model)
		writer.save()
		
		relevance_order = 0
		for index, row in relevant_features_df.iterrows():
			relevance_order = relevance_order + 1
			str_order = str(relevance_order)
			features_summary_df.set_value(current_gene, index, str_order)
			
	# Export the summary dataframe in an Excel file
	filename = './5_Data_Analysis/'+gene_set+'/Order_of_Features_Selected.xlsx'
	writer = ExcelWriter(filename,engine='openpyxl')
	try:
		writer.book = load_workbook(filename)
		writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
	except IOError:
		# if the file does not exist yet, I will create it
		pass
	features_summary_df.to_excel(writer,'M'+model)
	writer.save()


def summarize_r2(gene_set):

	"""
	The SUMMARIZE_R2 operation summarizes R2 and Adjusted R2 scores for each target gene in each regression model, storing them locally in a single Excel file.

	:param gene_set: the set of genes of interest to summarize
	
	Example::
	
		import genereg as gr
		gr.SummaryResults.summarize_r2(gene_set='DNA_REPAIR')
	"""

	
	# Define the models to summarize
	models = ['2','3','5']

	# Import the list of genes of interest and extract in a list the Gene Symbols of all the genes belonging to the current gene set
	EntrezConversion_df = pd.read_excel('./Genes_of_Interest.xlsx',sheetname='Sheet1',header=0,converters={'GENE_SYMBOL':str,'ENTREZ_GENE_ID':str,'GENE_SET':str})

	SYMs_current_pathway = []
	for index, row in EntrezConversion_df.iterrows():
		sym = row['GENE_SYMBOL']
		path = row['GENE_SET']
		if path == gene_set:
			SYMs_current_pathway.append(sym)

			
	# Create a dataframe to store the final summary about features selected and R2 scores, for each gene of interest
	summary_df = pd.DataFrame(index=SYMs_current_pathway, columns=['Selected Features (M2)','R2 (M2)','Adj.R2 (M2)','Selected Features (M3)','R2 (M3)','Adj.R2 (M3)','Selected Features (M5)','R2 (M5)','Adj.R2 (M5)'])

	for m in models:
		
		# Import the summary table for the current model
		current_df = pd.read_excel('./5_Data_Analysis/'+gene_set+'/Feature_Selection_and_Linear_Regression.xlsx',sheetname='M'+m,header=0)
		
		# Extract the useful information and store it the summary dataframe
		for index, row in current_df.iterrows():
			n_features = row['N° Features Selected']
			adj_r2 = row['Adj.R2']
			r2 = row['R2']
			
			summary_df.set_value(index,'Selected Features (M'+m+')',n_features)
			summary_df.set_value(index,'Adj.R2 (M'+m+')',adj_r2)
			summary_df.set_value(index,'R2 (M'+m+')',r2)

	# Export the summary dataframe in an Excel file
	summary_df = summary_df.sort_values(by=['Adj.R2 (M5)'], ascending=[False])
	writer = ExcelWriter('./5_Data_Analysis/'+gene_set+'/R2_and_Adj.R2_Scores.xlsx',engine='openpyxl')
	summary_df.to_excel(writer,'Sheet1')
	writer.save()
	

def best_genes(gene_set):

	"""
	The BEST_GENES operation collects the target genes with the best linear fit (Adjusted R2 >= 0.6) in the three regression models, storing them locally in a single Excel file.

	:param gene_set: the set of genes of interest to summarize
	
	Example::
	
		import genereg as gr
		gr.SummaryResults.best_genes(gene_set='DNA_REPAIR')		
	"""

	
	# Define the models to summarize
	models = ['2','3','5']

	# Import the list of genes of interest and extract in a list the Gene Symbols of all the genes belonging to the current gene set
	EntrezConversion_df = pd.read_excel('./Genes_of_Interest.xlsx',sheetname='Sheet1',header=0,converters={'GENE_SYMBOL':str,'ENTREZ_GENE_ID':str,'GENE_SET':str})

	SYMs_current_pathway = []
	for index, row in EntrezConversion_df.iterrows():
		sym = row['GENE_SYMBOL']
		path = row['GENE_SET']
		if path == gene_set:
			SYMs_current_pathway.append(sym)
			
	for model in models:        
        
		# Import the summary table cointaining the value of the R2 for each model and for each gene of interest in the current gene set
		# and extract the list of "good" genes, the ones that have R2 >= 0.6 in the current model
		summary_r2_df = pd.read_excel('./5_Data_Analysis/'+gene_set+'/R2_and_Adj.R2_Scores.xlsx',sheetname='Sheet1',header=0)
		summary_r2_df = summary_r2_df.sort_values(by=['Adj.R2 (M'+model+')'], ascending=[False])

		good_genes = []
		for index, row in summary_r2_df.iterrows():
			current_model_r2 = row['Adj.R2 (M'+model+')']
			if current_model_r2 >= 0.6:
				good_genes.append(index)

		# Create a dataframe to the store the results, indexed by the "good" genes and a progressive number for each significant feature extracted during the regression process
		num_features = []
		for i in list(range(1,20)):
			num_features.append(i)

		# Cartesian product to generate tuples for multi-indexing
		import itertools
		tuples = []
		for i in itertools.product(good_genes,num_features):
			tuples.append(i)

		# Set the multiple indexes to be used in the dataframe
		index = pd.MultiIndex.from_tuples(tuples, names=['GENE', '#'])

		# Create the dataframe and initialize the empty cells as empty strings
		final_summary_df = pd.DataFrame('', index = index, columns = ['Significant Feature','Adj.R2','Regression Coefficient','Feature Description']) 


		# Fill the dictionary

		# Adjusted R2
		for current_gene in good_genes:
			r2 = summary_r2_df.get_value(current_gene,'Adj.R2 (M'+model+')')
			final_summary_df.loc[(current_gene, 1),'Adj.R2'] = r2
			
		# Features
		for current_gene in good_genes:
			
			# Import the table containing the significant features extracted for the current gene
			gene_ID = EntrezConversion_df.loc[EntrezConversion_df['GENE_SYMBOL'] == current_gene, 'ENTREZ_GENE_ID'].iloc[0]
			features_df = pd.read_excel('./5_Data_Analysis/'+gene_set+'/Relevant_Features-Gene_'+gene_ID+'_['+current_gene+'].xlsx',sheetname='M'+model,header=0)
			
			feature_counter = 1
			for index, row in features_df.iterrows():
				coeff = row['Regression Coefficient']
				descr = row['Feature Description']
				final_summary_df.loc[(current_gene, feature_counter),'Significant Feature'] = index
				final_summary_df.loc[(current_gene, feature_counter),'Regression Coefficient'] = coeff
				final_summary_df.loc[(current_gene, feature_counter),'Feature Description'] = descr  
				feature_counter = feature_counter + 1

		# Remove the empty rows in the dataframe
		for index, row in final_summary_df.iterrows():
			feat = row['Significant Feature']
			coeff = row['Regression Coefficient']
			descr = row['Feature Description']
			if (feat == '') & (coeff == '') & (descr == ''):
				final_summary_df.drop(index, inplace=True)
				
		# Export the summary dataframe in an Excel file
		filename = './5_Data_Analysis/'+gene_set+'/Best_Genes.xlsx'
		writer = ExcelWriter(filename,engine='openpyxl')
		try:
			writer.book = load_workbook(filename)
			writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
		except IOError:
			# if the file does not exist yet, I will create it
			pass
		final_summary_df.to_excel(writer,'M'+model)
		writer.save()